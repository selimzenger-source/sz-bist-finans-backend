# -*- coding: utf-8 -*-
"""
Bilanço xlsx parser — Fintables/İş Yatırım formatı Excel dosyasından
finansal tablo çıkarımı. KAP parse başarısız/eksik kaldığında manuel
düzeltme için kullanılır.

Format (3 sayfa: Bilanço / Gelir Tablosu / Nakit Akış):
  - Satır 2: dönem kodları (B sütunundan itibaren): 202603, 202512, ...
    202603 = 2026/03 = 2026-Q1, 202606 = 2026-Q2, 202612 = 2026-Q4
  - Satır 3+: A sütunu = kalem adı (Türkçe), B sütunu = en güncel dönem değeri
  - Label'lar sektöre göre değişir (TOPLAM VARLIKLAR / VARLIKLAR TOPLAMI / AKTİF TOPLAMI)

Sektör-bağımsız: sanayi, banka, sigorta, holding, faktoring, aracı kurum, GMYO.
"""
import logging
import re

logger = logging.getLogger(__name__)


def _norm(s: str) -> str:
    """Türkçe → ASCII upper, match için."""
    if not s:
        return ""
    return (str(s).upper()
            .replace("İ", "I").replace("Ş", "S").replace("Ğ", "G")
            .replace("Ç", "C").replace("Ö", "O").replace("Ü", "U")
            .strip())


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) if v != 0 else None  # 0 = boş hücre kabul (xlsx başlık satırı 0 yazıyor)
    s = str(v).strip().replace(".", "").replace(",", ".")
    if not s or s in ("-", "—"):
        return None
    try:
        f = float(s)
        return f if f != 0 else None
    except ValueError:
        return None


def _period_label(code) -> str | None:
    """202603 → '2026-Q1'."""
    s = str(code).strip()
    m = re.match(r"^(\d{4})(\d{2})$", s)
    if not m:
        return None
    year, mm = int(m.group(1)), int(m.group(2))
    if mm not in (3, 6, 9, 12):
        return None
    return f"{year}-Q{mm // 3}"


# Kalem eşleme — sektör-bağımsız (exact-normalize match önce, sonra contains)
def _match_balance(label_norm: str) -> str | None:
    d = label_norm
    # total_assets
    if d in ("TOPLAM VARLIKLAR", "VARLIKLAR TOPLAMI", "AKTIF TOPLAMI", "TOPLAM AKTIFLER", "TOPLAM AKTIF"):
        return "total_assets"
    # total_equity — sanayi "TOPLAM OZKAYNAKLAR", sigorta "OZSERMAYE TOPLAMI",
    # banka tek başına "OZKAYNAKLAR" (TOPLAM yok). Sanayide de bare "OZKAYNAKLAR"
    # ara toplam = TOPLAM ile aynı değer → ilk eşleşme güvenli.
    if d in ("TOPLAM OZKAYNAKLAR", "OZKAYNAKLAR TOPLAMI", "OZSERMAYE TOPLAMI",
             "TOPLAM OZSERMAYE", "OZKAYNAK TOPLAMI", "OZKAYNAKLAR", "OZSERMAYE"):
        return "total_equity"
    # current / non-current
    if d in ("TOPLAM DONEN VARLIKLAR", "DONEN VARLIKLAR TOPLAMI"):
        return "current_assets"
    if d in ("TOPLAM DURAN VARLIKLAR", "DURAN VARLIKLAR TOPLAMI"):
        return "non_current_assets"
    if d in ("NAKIT VE NAKIT BENZERLERI",):
        return "cash_and_equivalents"
    return None


def _match_income(label_norm: str) -> str | None:
    d = label_norm
    # revenue (sektöre göre)
    if d in ("HASILAT", "SATIS GELIRLERI", "FAIZ GELIRLERI", "TOPLAM FAIZ GELIRI"):
        return "revenue"
    if "BRUT YAZILAN PRIM" in d or d == "TEKNIK GELIRLER":
        return "revenue"
    if d == "ESAS FAALIYET GELIRLERI":
        return "revenue"
    # gross profit
    if d in ("BRUT KAR (ZARAR)", "BRUT KAR", "BRUT KAR/ZARAR", "TICARI FAALIYETLERDEN BRUT KAR (ZARAR)"):
        return "gross_profit"
    # operating profit
    if d in ("ESAS FAALIYET KARI (ZARARI)", "ESAS FAALIYET KARI", "FAALIYET KARI (ZARARI)"):
        return "operating_profit"
    # net income
    if d in ("DONEM KARI (ZARARI)", "DONEM NET KARI (ZARARI)", "NET DONEM KARI (ZARARI)",
             "DONEM NET KARI", "NET DONEM KARI", "DONEM KARI/ZARARI"):
        return "net_income"
    return None


def parse_bilanco_xlsx(path: str, target_period: str | None = None) -> dict | None:
    """xlsx dosyasından EN GÜNCEL (veya target_period) dönemin finansallarını çıkar.

    Returns: {period, total_assets, total_equity, current_assets, non_current_assets,
              total_debt, cash_and_equivalents, revenue, gross_profit,
              operating_profit, net_income, _source_periods}
    Yoksa None.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        logger.warning("xlsx açılamadı %s: %s", path, e)
        return None

    out: dict = {}
    try:
        # Hangi sheet hangi tablo — isimle eşle (encoding-robust)
        sheets = {_norm(n): n for n in wb.sheetnames}
        bil_sheet = next((sheets[k] for k in sheets if "BILANCO" in k or "FINANSAL DURUM" in k), None)
        gel_sheet = next((sheets[k] for k in sheets if "GELIR" in k or "KAR VEYA ZARAR" in k), None)

        def _read_sheet(sheet_name, matcher):
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
            if len(rows) < 3:
                return {}, None
            # Dönem satırı: 2. satır (B sütunundan dönem kodları)
            period_row = rows[1]
            # Hedef sütun: target_period verilmişse onu bul, yoksa ilk dönem (B = index 1)
            col_periods = [(_period_label(period_row[c]), c) for c in range(1, len(period_row))
                           if _period_label(period_row[c])]
            if not col_periods:
                return {}, None
            if target_period:
                col = next((c for lbl, c in col_periods if lbl == target_period), col_periods[0][1])
                per = next((lbl for lbl, c in col_periods if c == col), col_periods[0][0])
            else:
                per, col = col_periods[0][0], col_periods[0][1]
            data = {}
            for row in rows[2:]:
                if not row or row[0] is None:
                    continue
                f = matcher(_norm(row[0]))
                if f and f not in data and col < len(row):
                    v = _to_float(row[col])
                    if v is not None:
                        data[f] = v
            return data, per

        period = None
        if bil_sheet:
            bal, period = _read_sheet(bil_sheet, _match_balance)
            out.update(bal)
        if gel_sheet:
            inc, p2 = _read_sheet(gel_sheet, _match_income)
            out.update(inc)
            period = period or p2

        if not period:
            wb.close()
            return None
        out["period"] = period
        wb.close()
        # En az bir anlamlı alan var mı?
        if not (out.get("total_assets") or out.get("revenue") or out.get("net_income")):
            return None
        return out
    except Exception as e:
        logger.warning("xlsx parse hata %s: %s", path, e)
        try:
            wb.close()
        except Exception:
            pass
        return None
